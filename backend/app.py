from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import io
import os
from datetime import datetime
import numpy as np
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

app = Flask(__name__)

def update_google_sheet(total_invested, total_market_value, total_unrealized_pl, total_day_pl, name_val, bank_val):
    creds_path = os.path.join(os.path.dirname(__file__), "update-sheet-497206-0ab70ed3823a.json")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    
    import gspread
    from google.oauth2.service_account import Credentials
    
    try:
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
    except Exception as e:
        raise Exception(f"Failed to authorize Google Sheets: {str(e)}")
        
    spreadsheet_id = "110by378Ie7TVMErdZ0NgOULDGXY43UKjS9udKykGxFk"
    sheet_name = "DAILY PL DETAILS"
    
    try:
        sh = client.open_by_key(spreadsheet_id)
    except Exception as e:
        raise Exception(
            f"Failed to open Google Sheet by ID: {str(e)}. "
            f"Please ensure you have shared the spreadsheet with the service account email: {creds.service_account_email}"
        )
        
    # Get or create worksheet
    try:
        ws = sh.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        try:
            ws = sh.add_worksheet(title=sheet_name, rows="1000", cols="20")
        except Exception as e:
            raise Exception(f"Failed to create sheet '{sheet_name}': {str(e)}")
            
    try:
        # Check date and session
        now = datetime.now()
        current_date = now.strftime("%d/%m/%Y")
        session = "Evening" if (now.hour >= 13 or now.hour < 9) else "Morning"
        
        if session != "Evening":
            print(f"Skipping Google Sheet update for Morning session report: {name_val} ({bank_val}).")
            return
            
        is_sudhakar_icici = (name_val.strip().upper() == "SUDHAKAR" and bank_val.strip().upper() == "ICICI")
        
        if is_sudhakar_icici:
                
            # --- Custom 11-Row Portfolio Dashboard Generation ---
            values = ws.get_all_values()
            
            # 1. Initialize default values based on the screenshot
            defaults = {
                ("SUDHAKAR", "ICICI DIRECT"): {
                    "profit": total_unrealized_pl,
                    "invested": total_invested,
                    "bank_balance": 379680,
                    "cmp_investment": total_market_value,
                    "profit_pct": total_unrealized_pl / total_invested if total_invested != 0 else 0,
                    "daily_profit": total_day_pl
                },
                ("NIRMALA", "ADITYA BIRLA"): {
                    "profit": 0.0,
                    "invested": 1.0,
                    "bank_balance": 1.0,
                    "cmp_investment": 1.0,
                    "profit_pct": 0.0,
                    "daily_profit": 1.0
                },
                ("SUDHAKAR", "ADITYA BIRLA"): {
                    "profit": 0.0,
                    "invested": 0.0,
                    "bank_balance": 0.0,
                    "cmp_investment": 0.0,
                    "profit_pct": 0.01,  # 1.00%
                    "daily_profit": 0.0
                },
                ("NIRMALA", "HDFC DIRECT"): {
                    "profit": 0.0,
                    "invested": 850.0,
                    "bank_balance": 5088.0,
                    "cmp_investment": 850.0,
                    "profit_pct": 0.0,
                    "daily_profit": 0.0
                },
                ("SUDHAKAR", "ZERODHA"): {
                    "profit": 0.0,
                    "invested": 0.0,
                    "bank_balance": 0.0,
                    "cmp_investment": 0.0,
                    "profit_pct": 0.0,
                    "daily_profit": 0.0
                },
                ("SUDHAKAR", "HDFCDIRECT"): {
                    "profit": 0.0,
                    "invested": 4036.0,
                    "bank_balance": 21151.0,
                    "cmp_investment": 4036.0,
                    "profit_pct": 0.0,
                    "daily_profit": 0.0
                }
            }
            
            # Helper to parse spreadsheet numeric formats
            def parse_num(val_str):
                if not val_str:
                    return 0.0
                s = str(val_str).strip().replace(',', '').replace('%', '')
                try:
                    return float(s)
                except:
                    return 0.0
            
            # 2. Bottom-up scanning to fetch the last known balances of all other portfolios
            current_values = defaults.copy()
            
            if values:
                for r in reversed(values):
                    if len(r) >= 12:
                        prev_name = str(r[2]).strip().upper()
                        prev_firm = str(r[3]).strip().upper()
                        key = (prev_name, prev_firm)
                        # Fetch the values for other combinations, leaving SUDHAKAR | ICICI DIRECT updated
                        if key in current_values and key != ("SUDHAKAR", "ICICI DIRECT"):
                            # Only populate if we haven't read a newer value yet
                            if current_values[key]["profit"] == defaults[key]["profit"] and current_values[key]["invested"] == defaults[key]["invested"] and current_values[key]["bank_balance"] == defaults[key]["bank_balance"]:
                                current_values[key] = {
                                    "profit": parse_num(r[6]),
                                    "invested": parse_num(r[7]),
                                    "bank_balance": parse_num(r[8]),
                                    "cmp_investment": parse_num(r[9]),
                                    "profit_pct": parse_num(r[10]) / 100.0 if '%' in str(r[10]) else parse_num(r[10]),
                                    "daily_profit": parse_num(r[11])
                                }
            
            # Determine starting row for formulas
            headers = [
                "", "S.No", "Name", "Broking Firm", "Today Profit", "Monthly Profit (Booked)",
                "Portfolio Overall Profit(UNREALISED PROFIT)", "Investment- Holding", 
                "Bank Balance ( Cash Balance)", "CMP Investment", "Profit%(unrealised)", "Daily Profit"
            ]
            
            if not values:
                start_row = 3
                rows_to_write = [headers]
                write_start_row = 2
            else:
                start_row = len(values) + 4
                rows_to_write = [
                    [""] * 12,
                    [""] * 12,
                    headers
                ]
                write_start_row = len(values) + 1
                
            # 3. Construct the 11 rows of data with exact formula offsets
            r1 = [current_date, "", "", "INVESTMENT", "", "", "", "", "", "", "", ""]
            r2 = [
                "", 1, "SUDHAKAR", "ICICI DIRECT", "", "", 
                round(current_values[("SUDHAKAR", "ICICI DIRECT")]["profit"]), 
                round(current_values[("SUDHAKAR", "ICICI DIRECT")]["invested"]), 
                round(current_values[("SUDHAKAR", "ICICI DIRECT")]["bank_balance"]), 
                f"=G{start_row+1}+H{start_row+1}", f"=IFERROR(G{start_row+1}/H{start_row+1},0)", 
                round(current_values[("SUDHAKAR", "ICICI DIRECT")]["daily_profit"])
            ]
            r3 = [
                "", 1, "NIRMALA", "ADITYA BIRLA", "", "", 
                round(current_values[("NIRMALA", "ADITYA BIRLA")]["profit"]), 
                round(current_values[("NIRMALA", "ADITYA BIRLA")]["invested"]), 
                round(current_values[("NIRMALA", "ADITYA BIRLA")]["bank_balance"]), 
                f"=G{start_row+2}+H{start_row+2}", f"=IFERROR(G{start_row+2}/H{start_row+2},0)", 
                round(current_values[("NIRMALA", "ADITYA BIRLA")]["daily_profit"])
            ]
            r4 = [
                "", "", "", "TOTAL", "", "", 
                f"=SUM(G{start_row+1}:G{start_row+2})", 
                f"=SUM(H{start_row+1}:H{start_row+2})", 
                f"=SUM(I{start_row+1}:I{start_row+2})", 
                f"=SUM(J{start_row+1}:J{start_row+2})", 
                f"=IFERROR(G{start_row+3}/H{start_row+3},0)", 
                f"=SUM(L{start_row+1}:L{start_row+2})"
            ]
            r_empty = [""] * 12
            r5 = ["", "", "", "", "", "", "TRADING", "", "", "", "", ""]
            r6 = [
                "", 1, "SUDHAKAR", "ADITYA BIRLA", "", "", 
                round(current_values[("SUDHAKAR", "ADITYA BIRLA")]["profit"]), 
                round(current_values[("SUDHAKAR", "ADITYA BIRLA")]["invested"]), 
                round(current_values[("SUDHAKAR", "ADITYA BIRLA")]["bank_balance"]), 
                f"=G{start_row+6}+H{start_row+6}", f"=IFERROR(G{start_row+6}/H{start_row+6},0)", 
                round(current_values[("SUDHAKAR", "ADITYA BIRLA")]["daily_profit"])
            ]
            r7 = [
                "", 2, "NIRMALA", "HDFC DIRECT", "", "", 
                round(current_values[("NIRMALA", "HDFC DIRECT")]["profit"]), 
                round(current_values[("NIRMALA", "HDFC DIRECT")]["invested"]), 
                round(current_values[("NIRMALA", "HDFC DIRECT")]["bank_balance"]), 
                f"=G{start_row+7}+H{start_row+7}", f"=IFERROR(G{start_row+7}/H{start_row+7},0)", 
                round(current_values[("NIRMALA", "HDFC DIRECT")]["daily_profit"])
            ]
            r8 = [
                "", "", "", "TOTAL", "", "", 
                f"=SUM(G{start_row+6}:G{start_row+7})", 
                f"=SUM(H{start_row+6}:H{start_row+7})", 
                f"=SUM(I{start_row+6}:I{start_row+7})", 
                f"=SUM(J{start_row+6}:J{start_row+7})", 
                f"=IFERROR(G{start_row+8}/H{start_row+8},0)", 
                f"=SUM(L{start_row+6}:L{start_row+7})"
            ]
            r_empty2 = [""] * 12
            r9 = [
                "", 1, "SUDHAKAR", "ZERODHA", "", "", 
                round(current_values[("SUDHAKAR", "ZERODHA")]["profit"]), 
                round(current_values[("SUDHAKAR", "ZERODHA")]["invested"]), 
                round(current_values[("SUDHAKAR", "ZERODHA")]["bank_balance"]), 
                f"=G{start_row+10}+H{start_row+10}", f"=IFERROR(G{start_row+10}/H{start_row+10},0)", 
                round(current_values[("SUDHAKAR", "ZERODHA")]["daily_profit"])
            ]
            r10 = [
                "", 2, "SUDHAKAR", "HDFCDIRECT", "", "", 
                round(current_values[("SUDHAKAR", "HDFCDIRECT")]["profit"]), 
                round(current_values[("SUDHAKAR", "HDFCDIRECT")]["invested"]), 
                round(current_values[("SUDHAKAR", "HDFCDIRECT")]["bank_balance"]), 
                f"=G{start_row+11}+H{start_row+11}", f"=IFERROR(G{start_row+11}/H{start_row+11},0)", 
                round(current_values[("SUDHAKAR", "HDFCDIRECT")]["daily_profit"])
            ]
            r11 = [
                "", "", "", "Total", "", "", 
                f"=G{start_row+3}+G{start_row+8}+G{start_row+10}+G{start_row+11}", 
                f"=H{start_row+3}+H{start_row+8}+H{start_row+10}+H{start_row+11}", 
                f"=I{start_row+3}+I{start_row+8}+I{start_row+10}+I{start_row+11}", 
                f"=J{start_row+3}+J{start_row+8}+J{start_row+10}+J{start_row+11}", 
                f"=IFERROR(G{start_row+12}/H{start_row+12},0)", 
                f"=L{start_row+3}+L{start_row+8}+L{start_row+10}+L{start_row+11}"
            ]
            
            rows_to_write.extend([r1, r2, r3, r4, r_empty, r5, r6, r7, r8, r_empty2, r9, r10, r11])
            ws.update(f"A{write_start_row}:L{write_start_row + len(rows_to_write) - 1}", rows_to_write, value_input_option='USER_ENTERED')
            
            try:
                # Format the written range to ensure all values are centered, middle-aligned, and wrapped
                ws.format(f"A{write_start_row}:L{write_start_row + len(rows_to_write) - 1}", {
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP"
                })
                
                # Specifically format the header row to remove bold formatting
                header_row_idx = write_start_row if not values else write_start_row + 2
                ws.format(f"A{header_row_idx}:L{header_row_idx}", {
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "wrapStrategy": "WRAP",
                    "textFormat": {
                        "bold": False
                    }
                })
                
                # Add double bottom border for header from B column to L
                ws.format(f"B{header_row_idx}:L{header_row_idx}", {
                    "borders": {
                        "bottom": {
                            "style": "DOUBLE"
                        }
                    }
                })
                
                # Add #f3f3f3 bg color for next row of header from B to L column
                ws.format(f"B{header_row_idx + 1}:L{header_row_idx + 1}", {
                    "backgroundColor": {
                        "red": 0.953,
                        "green": 0.953,
                        "blue": 0.953
                    }
                })
                
                # Add #f3f3f3 bg color for 3rd row from header from B to L column
                ws.format(f"B{header_row_idx + 3}:L{header_row_idx + 3}", {
                    "backgroundColor": {
                        "red": 0.953,
                        "green": 0.953,
                        "blue": 0.953
                    }
                })
                
                # Add #f9cb9c bg color and top/bottom bold (SOLID_MEDIUM) borders for 4th row from header (TOTAL row) from B to L
                ws.format(f"B{header_row_idx + 4}:L{header_row_idx + 4}", {
                    "backgroundColor": {
                        "red": 0.976,
                        "green": 0.796,
                        "blue": 0.612
                    },
                    "borders": {
                        "top": {
                            "style": "SOLID_MEDIUM"
                        },
                        "bottom": {
                            "style": "SOLID_MEDIUM"
                        }
                    }
                })
                
                # Add #93c47d bg color for 5th to 9th row from header from B to L column (TRADING section)
                ws.format(f"B{header_row_idx + 5}:L{header_row_idx + 9}", {
                    "backgroundColor": {
                        "red": 0.576,
                        "green": 0.769,
                        "blue": 0.490
                    }
                })
                
                # Add top and bottom bold border for 9th row from header (TRADING TOTAL) from B to L
                ws.format(f"B{header_row_idx + 9}:L{header_row_idx + 9}", {
                    "borders": {
                        "top": {
                            "style": "SOLID_MEDIUM"
                        },
                        "bottom": {
                            "style": "SOLID_MEDIUM"
                        }
                    }
                })
                
                # Add #ffe599 bg color for 10th to 12th row from header from B to L column (OTHERS section)
                ws.format(f"B{header_row_idx + 10}:L{header_row_idx + 12}", {
                    "backgroundColor": {
                        "red": 1.0,
                        "green": 0.898,
                        "blue": 0.600
                    }
                })
                
                # Add #deebf6 bg color for 13th row from header (Overall Total) from B to L
                ws.format(f"B{header_row_idx + 13}:L{header_row_idx + 13}", {
                    "backgroundColor": {
                        "red": 0.871,
                        "green": 0.922,
                        "blue": 0.965
                    }
                })
                # Add top and bottom lite (thin) border for 13th row from header (Overall Total) from B to L
                ws.format(f"B{header_row_idx + 13}:L{header_row_idx + 13}", {
                    "borders": {
                        "top": {
                            "style": "SOLID"
                        },
                        "bottom": {
                            "style": "SOLID"
                        }
                    }
                })
            except Exception as format_err:
                print(f"Non-fatal formatting error: {format_err}")
            
        else:
            # Standard single-row history logging for other accounts or morning sessions
            values = ws.get_all_values()
            headers = [
                "Date", "Session", "Name", "Bank", 
                "Total Invested", "Current Value", "Unrealized P&L", "Day P&L"
            ]
            if not values:
                ws.append_row(headers)
                
            row = [
                current_date,
                session,
                name_val,
                bank_val,
                round(total_invested, 2),
                round(total_market_value, 2),
                round(total_unrealized_pl, 2),
                round(total_day_pl, 2)
            ]
            ws.append_row(row, value_input_option='USER_ENTERED')
            
    except Exception as e:
        raise Exception(f"Error reading/writing to Google Sheet: {str(e)}")

# Hardcoded credentials for simplicity
HARDCODED_USERS = {
    "admin": {
        "password_hash": generate_password_hash("Admin@123"),
        "name": "System Admin",
        "role": "admin"
    }
}

# Enable CORS for React frontend
CORS(app, resources={r"/*": {"origins": "*"}})

def clean_numeric(val):
    if pd.isna(val) or val == '':
        return 0.0
    s = str(val).strip()
    # Handle parentheses for negative numbers: (10.22) -> -10.22
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    # Remove commas, spaces, and plus signs
    s = s.replace(',', '').replace(' ', '').replace('+', '')
    try:
        return float(s)
    except:
        return 0.0

@app.route('/process-excel', methods=['POST'])
def process_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    try:
        # Get dynamic title info early to determine style and parsing
        name_val = request.form.get('name', 'Gautham')
        bank_val = request.form.get('bank', 'HDFC')
        sync_to_sheets = request.form.get('sync_to_sheets', 'false').lower() == 'true'
        sheet_sync_success = False
        sheet_sync_error = None

        # Read input file
        df = None
        try:
            file.seek(0)
            if file.filename.endswith('.xls'):
                # Using the specific method requested by the user for .xls files
                df = pd.read_csv(file, sep="\t", engine='python', on_bad_lines='skip')
            elif file.filename.endswith('.xlsx'):
                df = pd.read_excel(file, engine='openpyxl')
            else:
                # General fallback (read CSV with column range to prevent dropping rows due to multi-column header/title rows)
                file.seek(0)
                df = pd.read_csv(file, sep=None, names=range(50), engine='python')
                df = df.dropna(how='all', axis=1)
        except Exception as e:
            # Final attempt if extension-based logic fails
            try:
                file.seek(0)
                df = pd.read_excel(file, engine='openpyxl')
            except Exception as e_xl:
                try:
                    file.seek(0)
                    df = pd.read_csv(file, sep=None, names=range(50), engine='python')
                    df = df.dropna(how='all', axis=1)
                except Exception as e2:
                    return jsonify({"error": f"Could not read file: {str(e2)}"}), 400

        if df is None or df.empty:
            return jsonify({"error": "File is empty or could not be parsed"}), 400
        
        # Clean column names (strip whitespace)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Auto-promote headers if they are located in the first few rows (density-based matching)
        candidates = [list(df.columns)]
        for idx in range(min(10, len(df))):
            candidates.append(list(df.iloc[idx]))
            
        header_keys = {
            'symbol', 'qty', 'quantity', 'avg price', 'average price', 'ltp', 'cmp', 
            'investment value', 'current value', 'unrealised p&l', 'unrealised p&l %', 
            'unrealized profit', 'company name', 'value at cost', 'value at market'
        }
        
        best_match_idx = -1
        max_matches = 0
        
        for c_idx, candidate in enumerate(candidates):
            matches = 0
            for val in candidate:
                val_str = str(val).strip().lower()
                if any(k in val_str for k in header_keys):
                    matches += 1
            if matches > max_matches:
                max_matches = matches
                best_match_idx = c_idx
                
        if best_match_idx > 0 and max_matches >= 3:
            actual_row_idx = best_match_idx - 1
            df.columns = [str(val).strip() for val in df.iloc[actual_row_idx].values]
            df = df.iloc[actual_row_idx + 1:].reset_index(drop=True)
            df.columns = [str(c).strip() for c in df.columns]
        
        # Qty value-shifting detection (HDFC anomaly)
        for col_name in df.columns:
            if col_name.lower().strip() == 'qty':
                qty_values = df[col_name]
                # Check if it has 0 non-empty values
                non_empty_count = qty_values.dropna().astype(str).str.strip().replace('', np.nan).dropna().count()
                if non_empty_count == 0:
                    # Find index of the column
                    col_idx = list(df.columns).index(col_name)
                    if col_idx + 1 < len(df.columns):
                        # Shift values from next column
                        df[col_name] = df.iloc[:, col_idx + 1]
                break
        
        # Check if the file matches the new column structure
        has_new_style = (bank_val == 'HDFC')
        
        # Mapping as per main.ipynb and user request
        if has_new_style:
            mapping = {
                'SYMBOL': 'SYMBOL',
                'QTY': 'QTY',
                'AVG PRICE': 'AVG PRICE',
                'LTP': 'LTP',
                'INVESTMENT VALUE': 'INVESTED',
                'CURRENT VALUE': 'CURRENT VALUE',
                'UNREALISED P&L': 'UNREALISED P&L',
                'UNREALISED P&L %': 'UNREALISED P&L %',
                "TODAY'S P&L": 'Day P&L',
                "TODAY'S CHG %": 'Day CHG %'
            }
        else:
            mapping = {
                'Company Name': 'SYMBOL',
                'Qty': 'Qty',
                'Average Cost Price': 'Avg Price',
                'Current Market Price': 'CMP',
                'Value At Cost': 'Invested',
                'Value At Market Price': 'Market Price',
                'Unrealized Profit/Loss': 'Unrealized P&L',
                'Unrealized Profit/Loss %': 'Unrealized P&L %',
                '% Change over prev close': '% Change'
            }
        
        # Filter and rename - ensure we don't crash if columns are missing
        processed_df = pd.DataFrame()
        for src, target in mapping.items():
            if src in df.columns:
                processed_df[target] = df[src]
            else:
                # Try case-insensitive match
                matches = [c for c in df.columns if c.lower() == src.lower()]
                if matches:
                    processed_df[target] = df[matches[0]]
                else:
                    # Robust fallback for partial matches/truncations
                    found = False
                    src_lower = src.lower()
                    for col in df.columns:
                        col_lower = col.lower()
                        if target in ['Company Name', 'SYMBOL'] and ('symbol' in col_lower or 'company' in col_lower):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Qty', 'QTY'] and col_lower.startswith('qty'):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Avg Price', 'AVG PRICE'] and ('avg' in col_lower or 'average' in col_lower):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['CMP', 'LTP'] and (col_lower == 'ltp' or 'current market' in col_lower):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Invested', 'INVESTED'] and ('investment' in col_lower or 'value at cost' in col_lower):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Market Price', 'CURRENT VALUE'] and ('current value' in col_lower or 'market price' in col_lower or col_lower == 'current'):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Unrealized P&L', 'UNREALISED P&L'] and 'unreal' in col_lower and '%' not in col_lower and 'pct' not in col_lower:
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['Unrealized P&L %', 'UNREALISED P&L %'] and 'unreal' in col_lower and ('%' in col_lower or 'pct' in col_lower or col_lower.endswith('.1')):
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target == 'Day P&L' and 'today' in col_lower and '%' not in col_lower and 'chg' not in col_lower:
                            processed_df[target] = df[col]
                            found = True
                            break
                        elif target in ['% Change', 'Day CHG %'] and ('change' in col_lower or 'chg' in col_lower or 'today\'s c' in col_lower):
                            processed_df[target] = df[col]
                            found = True
                            break
                    
                    if not found:
                        processed_df[target] = np.nan
        
        # Clean all numeric columns
        if has_new_style:
            numeric_cols = ['QTY', 'AVG PRICE', 'LTP', 'INVESTED', 'CURRENT VALUE', 'UNREALISED P&L', 'UNREALISED P&L %', 'Day P&L', 'Day CHG %']
        else:
            numeric_cols = ['Qty', 'Avg Price', 'CMP', 'Invested', 'Market Price', 'Unrealized P&L', 'Unrealized P&L %', 'Day P&L', '% Change']
            
        for col in numeric_cols:
            if col in processed_df.columns:
                processed_df[col] = processed_df[col].apply(clean_numeric)
        
        if not has_new_style:
            # Calculate Day P&L as per main.ipynb: (% Change * Market Price) / 100
            if '% Change' in processed_df.columns and 'Market Price' in processed_df.columns:
                processed_df['Day P&L'] = (processed_df['% Change'] * processed_df['Market Price']) / 100
            else:
                processed_df['Day P&L'] = 0.0

        # Round values as requested
        if has_new_style:
            # Round to integers (0 decimals)
            round_cols = ['INVESTED', 'CURRENT VALUE', 'UNREALISED P&L', 'UNREALISED P&L %', 'Day P&L']
            for col in round_cols:
                if col in processed_df.columns:
                    processed_df[col] = processed_df[col].round(0)
            
            # Round AVG PRICE and Day CHG % to 2 decimals
            decimals_cols = ['AVG PRICE', 'Day CHG %']
            for col in decimals_cols:
                if col in processed_df.columns:
                    processed_df[col] = processed_df[col].round(2)
        else:
            round_cols = ['Invested', 'Market Price', 'Unrealized P&L', 'Day P&L']
            for col in round_cols:
                if col in processed_df.columns:
                    processed_df[col] = processed_df[col].round(0)
        
        # Sort by Unrealized P&L % descending
        sort_col = 'UNREALISED P&L %' if has_new_style else 'Unrealized P&L %'
        if sort_col in processed_df.columns:
            processed_df = processed_df.sort_values(by=sort_col, ascending=False)
        
        # Reorder columns to match main.ipynb output
        if has_new_style:
            cols = ['SYMBOL', 'QTY', 'AVG PRICE', 'LTP', 'INVESTED', 'CURRENT VALUE', 'UNREALISED P&L', 'UNREALISED P&L %', 'Day P&L', 'Day CHG %']
        else:
            cols = ['SYMBOL', 'Qty', 'Avg Price', 'CMP', 'Invested', 'Market Price', 'Unrealized P&L', 'Unrealized P&L %', 'Day P&L', '% Change']
        # Ensure all columns exist before reordering
        final_cols = [c for c in cols if c in processed_df.columns]
        processed_df = processed_df[final_cols]
        
        # Calculate Totals
        if has_new_style:
            total_invested = processed_df['INVESTED'].sum() if 'INVESTED' in processed_df.columns else 0
            total_market_value = processed_df['CURRENT VALUE'].sum() if 'CURRENT VALUE' in processed_df.columns else 0
            total_unrealized_pl = processed_df['UNREALISED P&L'].sum() if 'UNREALISED P&L' in processed_df.columns else 0
            total_day_pl = processed_df['Day P&L'].sum() if 'Day P&L' in processed_df.columns else 0
            
            total_row = {
                'SYMBOL': np.nan,
                'QTY': np.nan,
                'AVG PRICE': np.nan,
                'LTP': np.nan,
                'INVESTED': total_invested,
                'CURRENT VALUE': total_market_value,
                'UNREALISED P&L': total_unrealized_pl,
                'UNREALISED P&L %': np.nan,
                'Day P&L': total_day_pl,
                'Day CHG %': np.nan
            }
        else:
            total_invested = processed_df['Invested'].sum() if 'Invested' in processed_df.columns else 0
            total_market_value = processed_df['Market Price'].sum() if 'Market Price' in processed_df.columns else 0
            total_unrealized_pl = processed_df['Unrealized P&L'].sum() if 'Unrealized P&L' in processed_df.columns else 0
            total_day_pl = processed_df['Day P&L'].sum() if 'Day P&L' in processed_df.columns else 0
            
            total_row = {
                'SYMBOL': 'TOTAL',
                'Qty': np.nan,
                'Avg Price': np.nan,
                'CMP': np.nan,
                'Invested': total_invested,
                'Market Price': total_market_value,
                'Unrealized P&L': total_unrealized_pl,
                'Unrealized P&L %': np.nan,
                'Day P&L': total_day_pl,
                '% Change': np.nan
            }
        
        processed_df = pd.concat([processed_df, pd.DataFrame([total_row])], ignore_index=True)
        
        # Sync daily totals to Google Sheet if requested in a background thread
        if sync_to_sheets:
            import threading
            
            def run_sync_background(t_inv, t_val, t_unreal, t_day, name, bank):
                try:
                    update_google_sheet(
                        total_invested=t_inv,
                        total_market_value=t_val,
                        total_unrealized_pl=t_unreal,
                        total_day_pl=t_day,
                        name_val=name,
                        bank_val=bank
                    )
                    print(f"Background Google Sheet sync completed successfully for {name} ({bank}).")
                except Exception as sync_err:
                    print(f"Background Google Sheet sync failed: {str(sync_err)}")
            
            try:
                # Start background thread
                thread = threading.Thread(
                    target=run_sync_background,
                    args=(
                        float(total_invested),
                        float(total_market_value),
                        float(total_unrealized_pl),
                        float(total_day_pl),
                        str(name_val),
                        str(bank_val)
                    )
                )
                thread.daemon = True
                thread.start()
                sheet_sync_success = True
            except Exception as thread_err:
                sheet_sync_error = str(thread_err)
                print(f"Failed to start background sync thread: {sheet_sync_error}")
        
        # Get dynamic title info (retrieved early at start of process_excel)
        
        # Determine Morning/Evening update
        now = datetime.now()
        # Treat hour >= 13 or hour < 9 as Evening (enables late night/early morning testing), and 9 to 12 as Morning
        update_status = "Evening" if (now.hour >= 13 or now.hour < 9) else "Morning"
        current_date = now.strftime("%d.%m.%Y")
        
        # Format title as per user request: ICICI (name) sir as on 13.05.2026 (Morning update)
        title = f"{bank_val} {name_val} sir as on {current_date} ({update_status} Update)"
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame([[title]]).to_excel(writer, index=False, header=False, startrow=0, startcol=0, sheet_name='Holdings')
            processed_df.to_excel(writer, index=False, startrow=1, sheet_name='Holdings')
            
            worksheet = writer.sheets['Holdings']
            # Auto-fit column widths with padding
            for col in worksheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Add padding of 2
                worksheet.column_dimensions[column_letter].width = max_length + 2

            # Format Title (Row 1) for both HDFC and standard styles
            title_cell = worksheet['A1']
            # Merge across all columns
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(processed_df.columns))
            title_cell.fill = PatternFill(start_color='FF262626', end_color='FF262626', fill_type='solid')
            title_cell.font = Font(color='FFFFFF', bold=True, size=12)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')


            
            # Center align all headers and values
            center_alignment = Alignment(horizontal='center', vertical='center')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_total_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
            next_row_fill = PatternFill(start_color='FFF3F3F3', end_color='FFF3F3F3', fill_type='solid')
            peach_fill = PatternFill(start_color='FFF9CB9C', end_color='FFF9CB9C', fill_type='solid')
            green_fill = PatternFill(start_color='FF93C47D', end_color='FF93C47D', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFE599', end_color='FFFFE599', fill_type='solid')
            blue_fill = PatternFill(start_color='FFDEEBF6', end_color='FFDEEBF6', fill_type='solid')
            bold_font = Font(bold=True)
            
            thin_side = Side(border_style="thin", color="000000")
            medium_side = Side(border_style="medium", color="000000")
            double_bottom_side = Side(border_style="double", color="000000")
            thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
            header_border_double = Border(top=thin_side, left=thin_side, right=thin_side, bottom=double_bottom_side)
            total_bold_border = Border(top=medium_side, left=thin_side, right=thin_side, bottom=medium_side)
            
            header_row_idx = 2
            total_row_idx = len(processed_df) + 2
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=len(processed_df.columns)), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if has_new_style:
                        if row_idx == 2 or row_idx == total_row_idx:
                            if row_idx == 2 and col_idx >= 2:
                                cell.border = header_border_double
                            else:
                                cell.border = thin_border
                            cell.alignment = header_alignment if row_idx == 2 else center_alignment
                            cell.fill = header_total_fill
                            if row_idx == total_row_idx:
                                cell.font = bold_font
                        elif row_idx > 2:
                            if row_idx == 6 and col_idx >= 2:
                                cell.border = total_bold_border
                                cell.fill = peach_fill
                            elif row_idx == 11 and col_idx >= 2:
                                cell.border = total_bold_border
                            elif row_idx == 13 and col_idx >= 2:
                                cell.border = thin_border
                            else:
                                cell.border = thin_border
                            cell.alignment = center_alignment
                            if (row_idx == 3 or row_idx == 5) and col_idx >= 2:
                                cell.fill = next_row_fill
                            elif 7 <= row_idx <= 10 and col_idx >= 2:
                                cell.fill = green_fill
                            elif 11 <= row_idx <= 12 and col_idx >= 2:
                                cell.fill = yellow_fill
                            elif row_idx == 13 and col_idx >= 2:
                                cell.fill = blue_fill
                    else:
                        if row_idx == 2 and col_idx >= 2:
                            cell.border = header_border_double
                        elif (row_idx == 6 or row_idx == 11) and col_idx >= 2:
                            cell.border = total_bold_border
                        elif row_idx == 13 and col_idx >= 2:
                            cell.border = thin_border
                        else:
                            cell.border = thin_border
                        if row_idx >= 2:
                            cell.alignment = header_alignment if row_idx == 2 else center_alignment
                            if row_idx == 2 or row_idx == total_row_idx:
                                cell.fill = header_total_fill
                                if row_idx == total_row_idx:
                                    cell.font = bold_font
                            elif (row_idx == 3 or row_idx == 5) and col_idx >= 2:
                                cell.fill = next_row_fill
                            elif row_idx == 6 and col_idx >= 2:
                                cell.fill = peach_fill
                            elif 7 <= row_idx <= 10 and col_idx >= 2:
                                cell.fill = green_fill
                            elif 11 <= row_idx <= 12 and col_idx >= 2:
                                cell.fill = yellow_fill
                            elif row_idx == 13 and col_idx >= 2:
                                cell.fill = blue_fill

            # Insert empty row between TOTAL (row 6) and TRADING section (row 7)
            worksheet.insert_rows(7)
            # Insert empty row as 10th from header (between TRADING TOTAL and ZERODHA)
            worksheet.insert_rows(12)
            
            # Set Row Heights
            worksheet.row_dimensions[1].height = 26 # Title is Row 1
            worksheet.row_dimensions[2].height = 38 # Header is Row 2
            for r_idx in range(3, worksheet.max_row + 1):
                worksheet.row_dimensions[r_idx].height = 18

            # Apply formatting
            if has_new_style:
                indian_format = '[>=10000000]##\,##\,##\,##0;[>=100000]##\,##\,##0;##,##0'
                integer_format = '#,##0'
                decimal_2_format = '0.00'
                
                # INVESTED, CURRENT VALUE, UNREALISED P&L, Day P&L -> indian_format
                for col_name in ['INVESTED', 'CURRENT VALUE', 'UNREALISED P&L', 'Day P&L']:
                    if col_name in processed_df.columns:
                        c_idx = processed_df.columns.get_loc(col_name) + 1
                        c_letter = get_column_letter(c_idx)
                        for r_idx in range(3, worksheet.max_row + 1): # Data starts at Row 3
                            cell = worksheet[f"{c_letter}{r_idx}"]
                            if cell.value is not None and cell.value != "":
                                cell.number_format = indian_format
                
                # UNREALISED P&L % -> integer_format
                if 'UNREALISED P&L %' in processed_df.columns:
                    c_idx = processed_df.columns.get_loc('UNREALISED P&L %') + 1
                    c_letter = get_column_letter(c_idx)
                    for r_idx in range(3, worksheet.max_row + 1):
                        cell = worksheet[f"{c_letter}{r_idx}"]
                        if cell.value is not None and cell.value != "":
                            cell.number_format = integer_format

                # AVG PRICE, Day CHG % -> decimal_2_format
                for col_name in ['AVG PRICE', 'Day CHG %']:
                    if col_name in processed_df.columns:
                        c_idx = processed_df.columns.get_loc(col_name) + 1
                        c_letter = get_column_letter(c_idx)
                        for r_idx in range(3, worksheet.max_row + 1):
                            cell = worksheet[f"{c_letter}{r_idx}"]
                            if cell.value is not None and cell.value != "":
                                cell.number_format = decimal_2_format
            else:
                indian_format = '[>=10000000]##\,##\,##\,##0;[>=100000]##\,##\,##0;##,##0'
                format_cols = ['Invested', 'Market Price', 'Unrealized P&L', 'Day P&L']
                for col_name in format_cols:
                    if col_name in processed_df.columns:
                        c_idx = processed_df.columns.get_loc(col_name) + 1
                        c_letter = get_column_letter(c_idx)
                        for r_idx in range(3, worksheet.max_row + 1):
                            worksheet[f"{c_letter}{r_idx}"].number_format = indian_format

            # Apply Conditional Formatting to 'UNREALISED P&L %' and 'Unrealized P&L %'
            cf_cols = ['UNREALISED P&L %'] if has_new_style else ['Unrealized P&L %']
            for col_name in cf_cols:
                if col_name in processed_df.columns:
                    col_idx = processed_df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    start_row = 3
                    end_row = 1 + len(processed_df)
                    if end_row >= start_row:
                        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                        
                        # Green-Yellow-Red Color Scale
                        rule = ColorScaleRule(
                            start_type='min', start_color='F8696B', # Red for low
                            mid_type='percentile', mid_value=50, mid_color='FFEB84', # Yellow for middle
                            end_type='max', end_color='63BE7B' # Green for high
                        )
                        worksheet.conditional_formatting.add(cell_range, rule)

        output.seek(0)
        
        # Format filename as per user request: name_bank_date.xlsx
        download_filename = f"{name_val}_{bank_val}_{current_date}.xlsx"
        
        resp = send_file(
            output,
            as_attachment=True,
            download_name=download_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp.headers['Access-Control-Expose-Headers'] = 'X-Sync-Success, X-Sync-Error'
        resp.headers['X-Sync-Success'] = 'true' if sheet_sync_success else 'false'
        if sheet_sync_error:
            resp.headers['X-Sync-Error'] = sheet_sync_error
        return resp
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- Authentication Routes ---

@app.route('/api/login/', methods=['POST'])
def login():
    try:
        data = request.json
        emp_code = data.get('employee_code', '').strip()
        password = data.get('password', '')
        
        user_data = HARDCODED_USERS.get(emp_code)
        if user_data:
            if check_password_hash(user_data["password_hash"], password):
                return jsonify({
                    'success': True,
                    'user': {
                        'employee_code': emp_code,
                        'name': user_data["name"],
                        'role': user_data["role"],
                        'is_initial_password': False
                    }
                })
        return jsonify({'success': False, 'message': 'Invalid employee code or password'}), 401
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=1501)
